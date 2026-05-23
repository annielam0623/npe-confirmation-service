"""
app/services/daily_report.py
Generates a 4-sheet Excel report for the next 3 days and emails it.
Sheets: Bus Tour | Tickets | Notes | Chats
Triggered at 11:59 PM LA time daily.
"""
import base64
import io
import logging
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.database import AsyncSessionLocal
from app.services.sendgrid import send_raw_email
from sqlalchemy import text

logger = logging.getLogger(__name__)
LA = ZoneInfo("America/Los_Angeles")

REPORT_RECIPIENTS = [
    "azhou@nationalparkexpress.com",
    "lxu@nationalparkexpress.com",
]

# ── Excel helpers ─────────────────────────────────────────────────────────────

def _make_header(ws, headers: list, color: str = "1A3A5C"):
    font  = Font(bold=True, color="FFFFFF", size=11)
    fill  = PatternFill("solid", fgColor=color)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = font
        cell.fill      = fill
        cell.alignment = align


def _write_rows(ws, rows, start_row: int = 2):
    align = Alignment(vertical="center", wrap_text=False)
    for ri, row in enumerate(rows, start_row):
        ws.row_dimensions[ri].height = 18
        for ci, val in enumerate(row, 1):
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            elif val is None:
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = align
            # Zebra stripe
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F8FF")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 52)


# ── Sheet builders ────────────────────────────────────────────────────────────

async def _build_bus_tour_sheet(ws, db, date_from, date_to):
    ws.title = "Bus Tour"
    headers = [
        "Order #", "Guest Name", "Phone", "Email", "Pax",
        "Tour Date", "Tour Type", "Pickup Time", "Pickup Location",
        "Conf Status", "Email Status", "SMS Status",
        "Turkey", "Veggie", "Beef",
        "MTLV Eligible", "MTLV Qty", "MTLV Ticket Status",
        "Agent", "Rezdy Status",
    ]
    _make_header(ws, headers, "2F7851")

    res = await db.execute(text("""
        SELECT
            b.order_number,
            TRIM(COALESCE(b.first_name,'') || ' ' || COALESCE(b.last_name,'')) AS guest_name,
            b.phone,
            b.customer_email,
            b.quantities,
            b.tour_date,
            b.tour_type,
            b.pickup_time,
            b.pickup_location,
            b.confirmation,
            b.email_status,
            b.sms_status,
            COALESCE(b.lunch_turkey, 0) AS turkey,
            COALESCE(b.lunch_veggie, 0) AS veggie,
            COALESCE(b.lunch_beef,  0) AS beef,
            b.mtlv_eligible,
            b.mtlv_qty,
            b.mtlv_ticket_status,
            b.agent_name,
            b.status
        FROM bookings b
        WHERE b.booking_type = 'bus_tour'
          AND b.tour_date BETWEEN :date_from AND :date_to
          AND b.status != 'cancelled'
        ORDER BY b.tour_date ASC, b.pickup_time ASC, b.last_name ASC
    """), {"date_from": date_from, "date_to": date_to})

    _write_rows(ws, res.all())
    _auto_width(ws)


async def _build_tickets_sheet(ws, db, date_from, date_to):
    ws.title = "Tickets"
    headers = [
        "Order #", "Guest Name", "Phone", "Email", "Pax",
        "Tour Date", "Tour Type", "Session Time", "Confirmation #",
        "Conf Status", "Email Status", "SMS Status",
        "MTLV Eligible", "MTLV Qty", "MTLV Ticket Status",
        "Agent", "Rezdy Status",
    ]
    _make_header(ws, headers, "1A3A5C")

    res = await db.execute(text("""
        SELECT
            b.order_number,
            TRIM(COALESCE(b.first_name,'') || ' ' || COALESCE(b.last_name,'')) AS guest_name,
            b.phone,
            b.customer_email,
            b.quantities,
            b.tour_date,
            b.tour_type,
            b.pickup_time        AS session_time,
            b.tt_number          AS confirmation_no,
            b.confirmation,
            b.email_status,
            b.sms_status,
            b.mtlv_eligible,
            b.mtlv_qty,
            b.mtlv_ticket_status,
            b.agent_name,
            b.status
        FROM bookings b
        WHERE b.booking_type = 'self_drive'
          AND b.tour_date BETWEEN :date_from AND :date_to
          AND b.status != 'cancelled'
        ORDER BY b.tour_date ASC, b.pickup_time ASC, b.last_name ASC
    """), {"date_from": date_from, "date_to": date_to})

    _write_rows(ws, res.all())
    _auto_width(ws)


async def _build_notes_sheet(ws, db, date_from, date_to):
    ws.title = "Notes"
    headers = [
        "Order #", "Tour Date", "Guest Name", "Note",
        "Author", "Created At",
    ]
    _make_header(ws, headers, "5F4AB7")

    res = await db.execute(text("""
        SELECT
            bn.order_number,
            b.tour_date,
            TRIM(COALESCE(b.first_name,'') || ' ' || COALESCE(b.last_name,'')) AS guest_name,
            bn.body,
            bn.author_username,
            bn.created_at AT TIME ZONE 'America/Los_Angeles' AS created_la
        FROM booking_notes bn
        LEFT JOIN bookings b ON b.order_number = bn.order_number
        WHERE bn.direction = 'staff_note'
          AND b.tour_date BETWEEN :date_from AND :date_to
        ORDER BY b.tour_date ASC, bn.created_at ASC
    """), {"date_from": date_from, "date_to": date_to})

    rows = []
    for r in res.all():
        created = r[5]
        if created and hasattr(created, "strftime"):
            created = created.strftime("%-m/%-d/%Y %-I:%M %p")
        rows.append((r[0], r[1], r[2], r[3], r[4], created))

    _write_rows(ws, rows)
    _auto_width(ws)


async def _build_chats_sheet(ws, db, date_from, date_to):
    ws.title = "Chats"
    headers = [
        "Order #", "Tour Date", "Guest Name", "Direction",
        "Message", "SMS Status", "Email Status", "Author", "Created At",
    ]
    _make_header(ws, headers, "BA7517")

    res = await db.execute(text("""
        SELECT
            bn.order_number,
            b.tour_date,
            TRIM(COALESCE(b.first_name,'') || ' ' || COALESCE(b.last_name,'')) AS guest_name,
            bn.direction,
            bn.body,
            bn.sms_status,
            bn.email_status,
            bn.author_username,
            bn.created_at AT TIME ZONE 'America/Los_Angeles' AS created_la
        FROM booking_notes bn
        LEFT JOIN bookings b ON b.order_number = bn.order_number
        WHERE bn.direction IN ('sms_out','email_out','sms_in','email_in','guest_reply')
          AND b.tour_date BETWEEN :date_from AND :date_to
        ORDER BY b.tour_date ASC, bn.created_at ASC
    """), {"date_from": date_from, "date_to": date_to})

    rows = []
    for r in res.all():
        created = r[8]
        if created and hasattr(created, "strftime"):
            created = created.strftime("%-m/%-d/%Y %-I:%M %p")
        rows.append((r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], created))

    _write_rows(ws, rows)
    _auto_width(ws)


# ── Main entry point ──────────────────────────────────────────────────────────

async def send_daily_report():
    """
    Build a 4-sheet Excel for the next 3 days and email it to REPORT_RECIPIENTS.
    Called at 11:59 PM LA time daily.
    """
    now_la    = datetime.now(LA)
    date_from = (now_la + timedelta(days=1)).date()
    date_to   = (now_la + timedelta(days=3)).date()
    label     = f"{date_from.strftime('%b %-d')} – {date_to.strftime('%b %-d, %Y')}"

    logger.info(f"[daily_report] Generating report for {date_from} → {date_to}")

    try:
        async with AsyncSessionLocal() as db:
            wb = openpyxl.Workbook()

            # Sheet 1: Bus Tour
            ws1 = wb.active
            await _build_bus_tour_sheet(ws1, db, date_from, date_to)

            # Sheet 2: Tickets
            ws2 = wb.create_sheet()
            await _build_tickets_sheet(ws2, db, date_from, date_to)

            # Sheet 3: Notes
            ws3 = wb.create_sheet()
            await _build_notes_sheet(ws3, db, date_from, date_to)

            # Sheet 4: Chats
            ws4 = wb.create_sheet()
            await _build_chats_sheet(ws4, db, date_from, date_to)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()

        # Base64 encode for SendGrid attachment
        encoded = base64.b64encode(xlsx_bytes).decode("utf-8")
        filename = f"NPE_Daily_Report_{date_from.strftime('%Y-%m-%d')}.xlsx"
        attachments = [{
            "content":     encoded,
            "type":        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename":    filename,
            "disposition": "attachment",
        }]

        subject = f"NPE Daily Operations Report — {label}"
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;">
          <h2 style="color:#1a3a5c;">NPE Daily Operations Report</h2>
          <p>Hi team,</p>
          <p>Please find attached the operations report for the next 3 days
             (<strong>{label}</strong>).</p>
          <p>The report includes 4 sheets:</p>
          <ul>
            <li><strong>Bus Tour</strong> — All bus tour bookings with MTLV &amp; lunch status</li>
            <li><strong>Tickets</strong> — Antelope Canyon ticket bookings</li>
            <li><strong>Notes</strong> — Internal staff notes</li>
            <li><strong>Chats</strong> — Guest SMS/Email conversations</li>
          </ul>
          <p style="color:#888;font-size:12px;">
            Generated automatically at 11:59 PM PT · National Park Express Operations
          </p>
        </div>
        """

        for recipient in REPORT_RECIPIENTS:
            await send_raw_email(
                to_email=recipient,
                to_name="NPE Team",
                subject=subject,
                html_body=html,
                attachments=attachments,
            )
            logger.info(f"[daily_report] Sent to {recipient}")

        logger.info(f"[daily_report] Done — {len(REPORT_RECIPIENTS)} recipients")

    except Exception as e:
        logger.error(f"[daily_report] Failed: {e}", exc_info=True)
        raise
